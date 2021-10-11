from openpyxl import load_workbook
from IPy import IP
import os


# Pergunta qual o nome do arquivo e Carrega
print(os.listdir('./'))
file = input("qual o nome do arquivo?")
caminho = f'./{file}'
arquivo_excel = load_workbook(caminho)

# Mostrar as abas do Excel e pergunta em qual está os dados
print(arquivo_excel.sheetnames)
xls_aba = input("qual o nome da aba?")
xls_data = arquivo_excel[xls_aba]

# Pergunta em quais colunas estão os dados Hostname e IP
letra_hostname = input("Qual a Letra da Coluna do Hostname?").upper()
letra_ip = input("Qual a Letra da Coluna do IP?").upper()
column_hostname = xls_data[letra_hostname]
column_ip = xls_data[letra_ip]

# Qual linha os Dados se iniciam
row_inicio_dados = input("Em qual linha começa os Dados?")

# Itera sobre os dados
file_hosts = open('./hosts.txt', 'w', encoding='utf-8')
for x in range(int(row_inicio_dados)-1, len(column_hostname)):
    hostname = column_hostname[x].value
    nat_ip = column_ip[x].value
    
    # Validade se os dados de ip são validos
    validade_ip = False
    try:
        IP(nat_ip)
        validade_ip = True
    except:
        validade_ip = False
    
    # se tudo validado escreve os dados    
    if(hostname and nat_ip and validade_ip):
        file_hosts.write(f'{nat_ip} {hostname}\n')

print('Arquivo Host Escrito!')        
file_hosts.close()
